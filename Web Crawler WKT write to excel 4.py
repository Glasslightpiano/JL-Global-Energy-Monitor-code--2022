# 四版省去excel合併至名為'Concat'欄位的前置作業
# 透過for迴圈加zip，同時讀取ProjectID和Wiki欄位資料
# 利用max_row自動帶入資料最大列數，不用手填while的筆數
# 需要手動調整的參數集中在此py檔的11~13列、18~19列，一目了然也方便好記
# 提高查看run成果的閱讀舒適度

from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests

file_route = 'read to write.xlsx'  # 輸入檔案路徑+名稱
sheet_name = 'sheet1'  # 輸入分頁名稱
head = 0  # 表頭列數，無表頭填0，一列表頭填1，兩列表頭填2，依此類推

wb = load_workbook(filename=file_route, read_only=False)  # read_only=False 才能讀 column
ws = wb[sheet_name]
COL = tuple(ws.columns)
ProjectID_cell = COL[0]  # 設定ProjectID欄位 excel A欄
Wiki_cell = COL[1]  # 設定Wiki欄位 excel B欄

row = head + 1
while row < ws.max_row:  # 自動帶入最大列數
    for ProjectID_output, Wiki_output in zip(ProjectID_cell, Wiki_cell):  # 合併迴圈 zip
        ID = ProjectID_output.value  # 顯示A欄數值
        URL = Wiki_output.value  # 顯示B欄數值
        # print(ID, ': ', URL)

        response = requests.get(URL)  # 爬蟲
        soup = BeautifulSoup(response.text, "html.parser")
        div_web = soup.find_all("div", class_="mapdata")\
                  and soup.find_all("div", id="map_google3_1")
        div_str = "".join(map(str, div_web))  # 爬蟲原始結果

        if ',"pos":[]' in div_str:  # 標示無座標資料
            resultI = (ID + ": No Coordinate!")
            print(resultI)
            ws.cell(row, 3).value = resultI[7:]  # 寫入excel檔C欄
            row += 1  # 換列

        elif ',"pos":[{' in div_str and div_str.count(',"pos":[{') == 1:  # Linestring
            L_split = div_str.split(',"pos":')
            for L_resp1 in L_split:
                if L_resp1.startswith('[{"lat":'):
                    L_resp2 = L_resp1.split(',"polygons":').pop(0)
                    L_str = L_resp2.replace('[', '').replace('{"lat":', '') \
                        .replace(',"lon":', ',').replace('},', '; ').replace('}]}]', '')
                    L_exchange1 = L_str.split('; ')
                    my_Linestring = ''
                    for L_exchange2 in L_exchange1:
                        L_exchange3 = L_exchange2.split(',')
                        L_exchange4 = L_exchange3[1] + ' ' + L_exchange3[0]
                        my_Linestring += str(L_exchange4 + ", ")
                    resultII = (ID + ": LINESTRING (" + my_Linestring[:-2] + ")")
                    print(resultII)
                    ws.cell(row, 3).value = resultII[7:]  # 寫入excel檔C欄
                    row += 1  # 換列

        elif ',"pos":[{' in div_str and div_str.count(',"pos":[{') > 1:  # Multilinestring
            M_splitI = div_str.split(',"pos":')
            Mmid_final = ''
            Mend_list = ''
            for M_splitII in M_splitI:
                if M_splitII.startswith('[{"lat":'):
                    if M_splitII.endswith(',"strokeWeight":"2"'):  # Multilinestring中段
                        Mmid_resp = M_splitII.split(',{"text":').pop(0)
                        Mmid_str = Mmid_resp.replace('[', '').replace('{"lat":', '') \
                            .replace(',"lon":', ',').replace('},', '; ').replace('}]}', '')
                        Mmid_exchange1 = Mmid_str.split('; ')
                        my_Mmid = ''
                        for Mmid_exchange2 in Mmid_exchange1:
                            Mmid_exchange3 = Mmid_exchange2.split(',')
                            Mmid_exchange4 = Mmid_exchange3[1] + ' ' + Mmid_exchange3[0]
                            my_Mmid += str(Mmid_exchange4 + ", ")
                        Mmid_list = "(" + my_Mmid[:-2] + "), "
                        for concat in Mmid_list:
                            Mmid_final += concat

                    elif M_splitII.endswith('</div>'):  # Multilinestring後段
                        Mend_resp = M_splitII.split('],"polygons":').pop(0)
                        Mend_str = Mend_resp.replace('[', '').replace('{"lat":', '') \
                            .replace(',"lon":', ',').replace('},', '; ').replace('}]}', '')
                        Mend_exchange1 = Mend_str.split('; ')
                        my_Mend = ''
                        for Mend_exchange2 in Mend_exchange1:
                            Mend_exchange3 = Mend_exchange2.split(',')
                            Mend_exchange4 = Mend_exchange3[1] + ' ' + Mend_exchange3[0]
                            my_Mend += str(Mend_exchange4 + ", ")
                        Mend_list = ("(" + my_Mend[:-2] + "))")
            resultIII = (ID + ": MULTILINESTRING (" + Mmid_final + Mend_list)
            print(resultIII)
            ws.cell(row, 3).value = resultIII[7:]  # 寫入excel檔C欄
            row += 1  # 換列

        else:  # 不符Global Energy Monitor原始碼規則
            resultIV = (ID + ": Not in the rule. Please check!")
            print(resultIV)
            ws.cell(row, 3).value = resultIV[7:]  # 寫入excel檔C欄
            row += 1  # 換列

wb.save(file_route)  # 儲存檔案
print("Finish!")
