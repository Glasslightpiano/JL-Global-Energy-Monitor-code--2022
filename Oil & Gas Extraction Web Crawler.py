# 前置作業：將要爬蟲資料的ID和WIKI欄位利用excel的concat公式合併至名為'Wiki_Concat'的欄位
# 需要調整的變數一：讀取excel檔名
# 需要調整的變數二：作業儲存用excel檔名和分頁編號

import pandas as pd
import requests
from bs4 import BeautifulSoup
import openpyxl

ds = pd.read_excel("Extraction_test.xlsx")   # 讀取excel檔
wb = openpyxl.load_workbook('Extraction_test.xlsx', data_only=True)  # 作業儲存用excel檔
sht = wb['sheet1']  # 作業儲存用excel檔分頁編號

row = 2
while row < 4:  # row < 任意數字，但要小於資料總筆數
    for URL in ds['Wiki_Concat']:   # 批次讀取網址
        response = requests.get(URL[9:])   # 爬蟲
        soup = BeautifulSoup(response.text, "html.parser")
        div_web = soup.find_all("div", class_="mapdata")   # 找原始碼<div> class="mapdata"
        div_str = ''.join(map(str, div_web))    # 轉換為string
        # print(URL[:9] + div_str)

        if 'lon' in div_str:  # 網址有坐標資料
            cut_first = div_str.split(',"link":"",')  # string拆解
            for cut_second in cut_first:
                if cut_second.startswith('"lat":'):
                    cut_third = cut_second.split(',"icon":')
                    for cut_fourth in cut_third:
                        if cut_fourth.startswith('"lat":'):
                            cut_fifth = cut_fourth.split(',')
                            # print(cut_fifth)
                            for coordinate in cut_fifth:  # string最後只留經緯度
                                if coordinate.startswith('"lat"'):  # 緯度
                                    latitude = coordinate.replace('"lat":', '')
                                    print(URL[:9] + " lat: " + latitude)
                                    sht.cell(row, 4).value = latitude  # 指定緯度儲存格

                                elif coordinate.startswith('"lon"'):  # 經度
                                    longitude = coordinate.replace('"lon":', '')
                                    print(URL[:9] + " lon: " + longitude)
                                    sht.cell(row, 5).value = longitude  # 指定經度儲存格
            row += 1  # excel換行

        else:  # 網址無座標資料
            print(URL[:9] + ": No coordinate!")
            sht.cell(row, 4).value = "--"  # 儲存格標示無坐標為"--"
            row += 1  # excel換行

wb.save('Extraction_test.xlsx')  # 儲存至指定excel檔
print("Finished!")
