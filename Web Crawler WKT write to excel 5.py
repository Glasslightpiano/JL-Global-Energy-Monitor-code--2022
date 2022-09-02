"""
五版修改重點：
一、變數命名改用駝峰式
二、使用 wet_format 的副程式執行爬蟲的原始座標轉換成 WKT 格式
     未使用副程式前這段相似的 code 在 Linestring 需執行一次； Multilinestring 需執行兩次
三、利用 re.split(' | '), ___ ，一次將兩個不同的 string 作長字串拆分點
四、盡可能合併語法，大幅減少變數的使用
五、整體語法簡化，少了十行左右

想改善但目前無法的部分：
結果匯出時必需明確指定 excel 儲存格，column 雖然都在第三欄，但 row 每輸出一筆結果需要換行再輸出下一筆
因此，無法省略 while 迴圈，輸入結果後要多加一行 row += 1，才能進行換行
"""

from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import re


def wkt_format(gcs):   # 原始座標轉換WKT格式的副程式
    substitute = gcs.replace('[', '').replace('{"lat":', '').replace(',"lon":', ',') \
        .replace('},', '; ').replace('}]}]', '').replace('}]}', '')
    # 倒數兩位的.replace('}]}]', '') for Linestring；.replace('}]}', '') for Multilinestring
    coordinate_split = substitute.split('; ')
    wkt_line = ''
    for turn1 in coordinate_split:
        turn2 = turn1.split(',')[1] + " " + turn1.split(',')[0]
        wkt_line += str(turn2 + ', ')
    return wkt_line   # *這行是副程式能正常運作重點*


fileRoute = 'read to write20220902.xlsx'   # 輸入檔案路徑+檔名
sheetName = 'sheet1'   # 輸入分頁名稱
head = 0  # 表頭列數，無表頭填0，一列表頭填1，兩列表頭填2，依此類推

workFile = load_workbook(filename=fileRoute, read_only=False)   # read_only=False 才能讀 column
workSheet = workFile[sheetName]
col = list(workSheet.columns)   # list或tuple都可以
cellProjectID = col[0]   # 設定ProjectID欄位excel A欄
cellWiki = col[1]   # 設定Wiki欄位excel B欄

row = head + 1
while row < workSheet.max_row:   # while迴圈停止於最大列數
    for project, wiki in zip(cellProjectID, cellWiki):   # 合併迴圈zip
        ID = project.value   # 以.value呈現欄位內的資料
        URL = wiki.value

        response = requests.get(URL)   # 爬蟲
        soup = BeautifulSoup(response.text, "html.parser")
        divWeb = soup.find_all("div", class_="mapdata") \
                 and soup.find_all("div", id="map_google3_1")
        divStr = "".join(map(str, divWeb))   # 爬蟲結果統一轉成str格式，批次存入divStr

        if ',"pos":[]' in divStr:   # 無座標資料
            resultI = (ID + ": No Coordinate!")
            print(resultI)
            workSheet.cell(row, 3).value = resultI[7:]   # 寫入excel C欄
            row += 1  # 換列

        elif ',"pos":[{' in divStr and divStr.count(',"pos":[{') == 1:   # Linestring
            lineSplit = re.split(',"pos":|,"polygons":', divStr)[1]   # re.split一次分兩個
            resultII = (ID + ": LINESTRING (" + wkt_format(lineSplit)[:-2] + ")")   # 副程式
            print(resultII)
            workSheet.cell(row, 3).value = resultII[7:]
            row += 1

        elif',"pos":[{' in divStr and divStr.count(',"pos":[{') > 1:   # Multilinestring
            multiMidCombine = ''
            multiEnd = ''
            multiSplit1 = re.split(',"pos":|,"polygons":', divStr)[1:]   # re.split一次分兩個
            for multiSplit2 in multiSplit1:
                if multiSplit2.endswith(',"strokeWeight":"2"'):   # Multilinestring中段
                    multiMidSplit = multiSplit2.split(',{"text":').pop(0)
                    multiMidList = ("(" + wkt_format(multiMidSplit)[:-2] + "), ")   # 副程式
                    for concat in multiMidList:
                        multiMidCombine += concat   # 中段的Linestring逐一合併

                elif multiSplit2.startswith('[{"lat":'):   # Multilinestring尾段
                    multiEnd = ("(" + wkt_format(multiSplit2)[:-2] + "))")   # 副程式

            resultIII = (ID + ": MULTILINESTRING (" + multiMidCombine + multiEnd)   # 總合併
            print(resultIII)
            workSheet.cell(row, 3).value = resultIII[7:]
            row += 1

        else:   # 不符Global Energy Monitor原始碼規則
            resultIV = (ID + ": Not in the rule. Please check!")
            print(resultIV)
            workSheet.cell(row, 3).value = resultIV[7:]
            row += 1

workFile.save(fileRoute)   # 儲存至excel
print("Finish!")
