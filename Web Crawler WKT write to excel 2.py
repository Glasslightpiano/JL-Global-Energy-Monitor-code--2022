# 二版更加簡約
# 以pop(0)直接選出List的第一個項目，省去一段for和if

import pandas as pd
import requests
from bs4 import BeautifulSoup
import openpyxl

ds = pd.read_excel("readtest.xlsx")  # 讀取excel檔
wb = openpyxl.load_workbook('writetest.xlsx', data_only=True)  # 作業儲存用excel檔
sht = wb['sheet1']  # 作業儲存用excel檔分頁編號

row = 2
while row < 5:  # row < 任意數字，但要小於資料總筆數
    for URL in ds['Concat']:  # 批次讀取網址
        response = requests.get(URL[5:])  # 爬蟲
        soup = BeautifulSoup(response.text, "html.parser")  # 搜尋原始碼class="mapdata" and id="map_google3_1"
        div_web = soup.find_all("div", class_="mapdata") and soup.find_all("div", id="map_google3_1")
        div_str = "".join(map(str, div_web))  # 轉換為string

        if ',"pos":[]' in div_str:  # 標示無坐標資料
            result_1 = (URL[:5] + ":No Coordinate!")
            print(result_1)
            sht.cell(row, 1).value = result_1  # 寫入儲存用excel檔
            row += 1  # excel換行

        elif ',"pos":[{' in div_str and div_str.count(',"pos":[{') == 1:  # 標示為Linestring
            L_startsplit = div_str.split(',"pos":')  # 去除非坐標訊息
            for L_1 in L_startsplit:
                if L_1.startswith('[{"lat":'):
                    L_2 = L_1.split(',"polygons":').pop(0)  # 從List中選出第一項
                    L_str = L_2.replace('[', '').replace('{"lat":', '').replace(',"lon":', ',') \
                        .replace('},', '; ').replace('}]}]', '')
                    Lturn_1 = L_str.split('; ')
                    my_Linestring = ''
                    for Lturn_2 in Lturn_1:  # 經緯度交換
                        Lturn_3 = Lturn_2.split(',')
                        Lturn_4 = Lturn_3[1] + ' ' + Lturn_3[0]
                        my_Linestring += str(Lturn_4 + ", ")  # 整合WKT格式
                    result_2 = (URL[:5] + ":LINESTRING (" + my_Linestring[:-2] + ")")
                    print(result_2)
                    sht.cell(row, 1).value = result_2  # 寫入儲存用excel檔
            row += 1  # excel換行

        elif ',"pos":[{' in div_str and div_str.count(',"pos":[{') > 1:  # 標示為Multilinestring
            M_step1 = div_str.split(',"pos":')  # 去除前段非坐標訊息
            Mmid_final = ''  # 整合最後一段以外的Linestring預留
            Mend_list = ''  # 為了加入最後一段的Linestring預留
            for M_step2 in M_step1:
                if M_step2.startswith('[{"lat":'):
                    if M_step2.endswith(',"strokeWeight":"2"'):  # 去除中段非坐標訊息
                        M_step3_mid = M_step2.split(',{"text":').pop(0)  # 從List中選出第一項
                        M_str_mid = M_step3_mid.replace('[', '').replace('{"lat":', '') \
                            .replace(',"lon":', ',').replace('},', '; ').replace('}]}', '')
                        M_turn1_mid = M_str_mid.split('; ')
                        my_Mmid = ''
                        for M_turn2_mid in M_turn1_mid:  # 經緯度交換
                            M_turn3_mid = M_turn2_mid.split(',')
                            M_turn4_mid = M_turn3_mid[1] + ' ' + M_turn3_mid[0]
                            my_Mmid += str(M_turn4_mid + ", ")
                        Mmid_list = "(" + my_Mmid[:-2] + "), "  # 整合WKT格式
                        # print(Mmid_list)
                        for concat in Mmid_list:  # 最後一段以外的Linestring整併
                            Mmid_final += concat

                    elif M_step2.endswith('</div>'):  # 去除後段非坐標訊息
                        M_step3_end = M_step2.split('],"polygons":').pop(0)  # 從List中選出第一項
                        M_str_end = M_step3_end.replace('[', '').replace('{"lat":', '') \
                            .replace(',"lon":', ',').replace('},', '; ').replace('}]}', '')
                        M_turn1_end = M_str_end.split('; ')
                        my_Mend = ''
                        for M_turn2_end in M_turn1_end:  # 經緯度交換
                            M_turn3_end = M_turn2_end.split(',')
                            M_turn4_end = M_turn3_end[1] + ' ' + M_turn3_end[0]
                            my_Mend += str(M_turn4_end + ", ")
                        Mend_list = ("(" + my_Mend[:-2] + "))")  # 整合WKT格式
                        # print(Mend_list)
            result_3 = (URL[:5] + ":MULTILINESTRING (" + Mmid_final + Mend_list)  # 加最後一段Linestring
            print(result_3)
            sht.cell(row, 1).value = result_3  # 寫入儲存用excel檔
            row += 1  # excel換行

        else:  # 不符Global Energy Monitor原始碼規則
            result_4 = (URL[:5] + ": Not in the rule. Please check!")
            print(result_4)
            sht.cell(row, 1).value = result_4  # 寫入儲存用excel檔
            row += 1  # excel換行

wb.save('writetest.xlsx')  # 儲存至指定excel檔
print("Finish!")
