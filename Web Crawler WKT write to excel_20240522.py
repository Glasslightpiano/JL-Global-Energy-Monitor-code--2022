"""
20240522改版重點：
一、爬蟲、轉 WKT格式、判斷坐標字串長度、寫入 excel或 csv檔於這支.py檔完成
二、使用 wet_format 的副程式(function)執行爬蟲的原始座標轉換成 WKT格式
    使用副程式前這段相似的 code在 Linestring需執行一次；Multilinestring需執行兩次

前置作業準備：
一、相同作業目錄資料夾下要有以下作業檔：
   (1)包含 Line屬性資料的 excel檔，A欄填 ID；B欄填網址。不需要表頭 "head = 0 "。
   (2)這支.py檔
   (3)一份完全空白的 csv檔，用來記錄坐標字串長度大於 32762的資料。

作業成果：
一、所有的結果紀錄都會寫回 excel檔的 C欄：
    (1)寫入 WKT Linestring坐標
    (2)寫入 WKT Multilinestring坐標
    (3)該網址提供的 WKT坐標字串太長，無法寫入 excel： length too long!
    (4)該網址未提供坐標資料： No Coordinate!
    (5)該網址的網頁原始碼不符 Global Energy Monitor爬蟲規則： Not in the rule. Please check!
二、完成寫入的 excel檔篩選無坐標欄位檢查，檢查無誤後可另存為.csv格式。
三、WKT坐標字串太長會自動寫入空白 csv檔，請用 Notepad++開啟，複製 WKT坐標後貼在第二點另存的.csv檔的同名 ID後方
四、另存的.csv檔 C欄都有 WKT坐標後可使用QGIS讀取
"""

from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import re
import csv


def wkt_format(gcs):  # 原始座標轉換WKT格式的副程式
    substitute = gcs.replace('[', '').replace('{"lat":', '').replace(',"lon":', ',') \
        .replace('},', '; ').replace('}]}]', '').replace('}]}', '')
    # 倒數兩位的.replace('}]}]', '') for Linestring；.replace('}]}', '') for Multilinestring
    coordinate_split = substitute.split('; ')
    wkt_line = ''
    for turn1 in coordinate_split:
        turn2 = turn1.split(',')[1] + " " + turn1.split(',')[0]
        wkt_line += str(turn2 + ', ')
    return wkt_line  # *這行是副程式能正常運作重點*


# 讀取、匯出檔案路徑及名稱設定
fileRoute = 'Global_Gas_Pipeline_test.xlsx'  # 輸入檔案路徑+檔名
sheetName = 'sheet1'  # 輸入excel分頁名稱
head = 0  # 表頭列數，無表頭填0，一列表頭填1，兩列表頭填2，依此類推
writeRoute = open("Gas_Pipeline_too_long.csv", 'w')  # WKT坐標字串大於等於32762則寫入csv檔
wr = csv.writer(writeRoute)

# load_workbook套件讀取 excel欄位設定
workFile = load_workbook(filename=fileRoute, read_only=False)  # read_only=False 才能讀 column
workSheet = workFile[sheetName]
col = list(workSheet.columns)  # list或tuple都可以
cellProjectID = col[0]  # 設定ProjectID欄位excel A欄
cellWiki = col[1]  # 設定Wiki欄位excel B欄

# 開始執行作業
row = head + 1
for project, wiki in zip(cellProjectID, cellWiki):  # 合併迴圈zip
    ID = project.value  # 以.value呈現欄位內的資料
    URL = wiki.value

    response = requests.get(URL)  # 爬蟲
    soup = BeautifulSoup(response.text, "html.parser")
    divWeb = soup.find_all("div", class_="mapdata") \
            and soup.find_all("div", id="map_google3_1")
    divStr = "".join(map(str, divWeb))  # 爬蟲結果統一轉成str格式，批次存入divStr

    if ',"pos":[]' in divStr:  # 無座標資料
        resultI = (ID + ": No Coordinate!")
        print(resultI)
        workSheet.cell(row, 3).value = resultI[7:]   # 寫入excel C欄
        row += 1  # 換列

    elif ',"pos":[{' in divStr and divStr.count(',"pos":[{') == 1:  # Linestring
        lineSplit = re.split(',"pos":|,"polygons":', divStr)[1]  # re.split將兩個不同的 string作長字串拆分點
        resultII = (ID + ": LINESTRING (" + wkt_format(lineSplit)[:-2] + ")")  # 副程式
        if len(resultII) < 32762:  # Linestring長度小於32762直接寫入excel
            print(resultII)
            workSheet.cell(row, 3).value = resultII[7:]
            row += 1
        elif len(resultII) >= 32762:   # Linestring長度大於等於32762改寫入csv
            print(ID + ": Linestring length too long!")
            wr.writerow([resultII])
            workSheet.cell(row, 3).value = "length too long!"  # excel註記 length too long
            row += 1

    elif ',"pos":[{' in divStr and divStr.count(',"pos":[{') > 1:  # Multilinestring
        multiMidCombine = ''
        multiEnd = ''
        multiSplit1 = re.split(',"pos":|,"polygons":', divStr)[1:]  # re.split將兩個不同的 string作長字串拆分點
        for multiSplit2 in multiSplit1:
            if multiSplit2.endswith(',"strokeWeight":"2"'):  # Multilinestring中段
                multiMidSplit = multiSplit2.split(',{"text":').pop(0)
                multiMidList = ("(" + wkt_format(multiMidSplit)[:-2] + "), ")  # 副程式
                for concat in multiMidList:
                    multiMidCombine += concat  # 中段的Linestring逐一合併

            elif multiSplit2.startswith('[{"lat":'):  # Multilinestring尾段
                multiEnd = ("(" + wkt_format(multiSplit2)[:-2] + "))")  # 副程式

        resultIII = (ID + ": MULTILINESTRING (" + multiMidCombine + multiEnd)  # 總合併
        if len(resultIII) < 32762:  # Multilinestring長度小於32762直接寫入excel
            print(resultIII)
            workSheet.cell(row, 3).value = resultIII[7:]
            row += 1
        elif len(resultIII) >= 32762:   # Multilinestring長度大於等於32762改寫入csv
            print(ID + ": Multilinestring length too long!")
            wr.writerow([resultIII])
            workSheet.cell(row, 3).value = "length too long!"  # excel註記 length too long
            row += 1

    else:  # 不符Global Energy Monitor原始碼規則
        resultIV = (ID + ": Not in the rule. Please check!")
        print(resultIV)
        workSheet.cell(row, 3).value = resultIV[7:]
        row += 1

workFile.save(fileRoute)   # 儲存至excel
print("Finish!")
