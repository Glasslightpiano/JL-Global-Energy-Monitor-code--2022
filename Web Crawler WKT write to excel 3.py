# 僅使用openpyxl套件完成在相同excel檔讀寫
# pandas套件似乎只能寫入list (待確認)
# 加強命名辨識度
# 長句換行

from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests

file_route = 'read to write.xlsx'  # 輸入檔案路徑+名稱

wb = load_workbook(filename=file_route, read_only=False)  # read_only=False 才能讀 column
ws = wb['sheet1']  # 輸入分頁名稱

row = 1
while row < 5:  # row < 任意數字，但要小於資料總筆數
    COL = tuple(ws.columns)
    column_designate = COL[2]  # 指定C欄
    for cell in column_designate:
        URL = cell.value  # 讀取C欄，格式為string
        response = requests.get(URL[5:])  # 爬蟲
        soup = BeautifulSoup(response.text, "html.parser")
        div_web = soup.find_all("div", class_="mapdata")\
                  and soup.find_all("div", id="map_google3_1")
        div_str = "".join(map(str, div_web))  # 爬蟲原始結果

        if ',"pos":[]' in div_str:  # 標示無座標資料
            resultI = (URL[:5] + ":No Coordinate!")
            print(resultI)
            ws.cell(row, 4).value = resultI[6:]  # 寫入excel檔D欄
            row += 1  # 換列

        elif ',"pos":[{' in div_str and div_str.count(',"pos":[{') == 1:  # Linestring
            L_split = div_str.split(',"pos":')
            for L_resp1 in L_split:
                if L_resp1.startswith('[{"lat":'):
                    L_resp2 = L_resp1.split(',"polygons":').pop(0)
                    L_str = L_resp2.replace('[', '').replace('{"lat":', '')\
                        .replace(',"lon":', ',').replace('},', '; ').replace('}]}]', '')
                    L_exchange1 = L_str.split('; ')
                    my_Linestring = ''
                    for L_exchange2 in L_exchange1:
                        L_exchange3 = L_exchange2.split(',')
                        L_exchange4 = L_exchange3[1] + ' ' + L_exchange3[0]
                        my_Linestring += str(L_exchange4 + ", ")
                    resultII = (URL[:5] + ":LINESTRING (" + my_Linestring[:-2] + ")")
                    print(resultII)
                    ws.cell(row, 4).value = resultII[6:]  # 寫入excel檔D欄
                    row += 1  # 換列

        elif ',"pos":[{' in div_str and div_str.count(',"pos":[{') > 1:  # Multilinestring
            M_splitI = div_str.split(',"pos":')
            Mmid_final = ''
            Mend_list = ''
            for M_splitII in M_splitI:
                if M_splitII.startswith('[{"lat":'):
                    if M_splitII.endswith(',"strokeWeight":"2"'):  # Multilinestring中段
                        Mmid_resp = M_splitII.split(',{"text":').pop(0)
                        Mmid_str = Mmid_resp.replace('[', '').replace('{"lat":', '')\
                            .replace(',"lon":', ',').replace('},', '; ').replace('}]}', '')
                        Mmid_exchange1 = Mmid_str.split('; ')
                        my_Mmid = ''
                        for Mmid_exchange2 in Mmid_exchange1:
                            Mmid_exchange3 = Mmid_exchange2.split(',')
                            Mmid_exchange4 = Mmid_exchange3[1] + ' ' + Mmid_exchange3[0]
                            my_Mmid += str(Mmid_exchange4 + ", ")
                        Mmid_list = "(" + my_Mmid[:-2] + "), "
                        for concat in Mmid_list:
                            Mmid_final += concat

                    elif M_splitII.endswith('</div>'):  # Multilinestring後段
                        Mend_resp = M_splitII.split('],"polygons":').pop(0)
                        Mend_str = Mend_resp.replace('[', '').replace('{"lat":', '')\
                            .replace(',"lon":', ',').replace('},', '; ').replace('}]}', '')
                        Mend_exchange1 = Mend_str.split('; ')
                        my_Mend = ''
                        for Mend_exchange2 in Mend_exchange1:
                            Mend_exchange3 = Mend_exchange2.split(',')
                            Mend_exchange4 = Mend_exchange3[1] + ' ' + Mend_exchange3[0]
                            my_Mend += str(Mend_exchange4 + ", ")
                        Mend_list = ("(" + my_Mend[:-2] + "))")
            resultIII = (URL[:5] + ":MULTILINESTRING (" + Mmid_final + Mend_list)
            print(resultIII)
            ws.cell(row, 4).value = resultIII[6:]  # 寫入excel檔D欄
            row += 1  # 換列

        else:
            resultIV = (URL[:5] + ":Not in the rule. Please check!")
            print(resultIV)
            ws.cell(row, 4).value = resultIV[6:]  # 寫入excel檔D欄
            row += 1  # 換列

wb.save(file_route)  # 儲存檔案
print("Finish!")
