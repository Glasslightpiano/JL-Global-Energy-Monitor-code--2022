from openpyxl import load_workbook

wb = load_workbook("GFIT_readtest.xlsx", data_only=True)  # 讀取excel檔
sht = wb['sheet1']  # 選擇工作表

cellRange = sht['B']  # 選擇坐標欄位
row = 1
while row < 10:  # row < 任意數字，但要小於資料總筆數
    for line in cellRange:  # 列出坐標
        origional = line.value
        # print(origional)

        if ';' in origional:  # 篩選出屬於Multilinestring的資料
            Mspl_first = origional.split(';')

            Mcnt_fifth = ''
            for Mspl_second in Mspl_first:
                Mspl_third = Mspl_second.split(':')  # 去除原始資料不同點位以':'分隔
                Mcnt_second = ''
                for Mspl_fourth in Mspl_third:
                    Mspl_fifth = Mspl_fourth.split(',')  # 原始資料相同點位經緯度以','分隔
                    Mcnt_first = Mspl_fifth[1] + ' ' + Mspl_fifth[0]  # 資料重組為WKT格式
                    Mcnt_second += str(Mcnt_first + ", ")
                Mcnt_third = ("(" + Mcnt_second[:-2] + "), ")  # 合併為Linestring格式
                for Mcnt_fourth in Mcnt_third:
                    Mcnt_fifth += Mcnt_fourth
            Mresult = ("MULTILINESTRING (" + Mcnt_fifth[:-2] + ")")  # 合併為Multilinestring格式
            print(Mresult)
            sht.cell(row, 3).value = Mresult  # 寫入excel儲存格
            row += 1  # excel換行

        else:  # 字串中沒有';'屬於Linestring的資料
            Lspl_first = origional.split(':')  # 去除原始資料不同點位以':'分隔

            Lcnt_second = ''
            for Lspl_second in Lspl_first:
                Lspl_third = Lspl_second.split(',')  # 原始資料相同點位經緯度以','分隔
                Lcnt_first = Lspl_third[1] + ' ' + Lspl_third[0]  # 資料重組為WKT格式
                Lcnt_second += str(Lcnt_first + ", ")
            Lresult = ("LINESTRING (" + Lcnt_second[:-2] + ")")  # 合併為Linestring格式
            print(Lresult)
            sht.cell(row, 3).value = Lresult  # 寫入excel儲存格
            row += 1  # excel換行

wb.save('GFIT_readtest.xlsx')  # 儲存至指定excel檔
print("Finish!")
